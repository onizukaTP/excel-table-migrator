import csv
import math
import re
import datetime
import logging
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

def clean_cell_value(val):
    """Normalize Excel/CSV cell values:
    - None -> None
    - NaN -> None
    - str -> stripped, empty -> None
    - datetime/date -> ISO string
    - int/float/bool -> preserved
    """
    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.isoformat()
    if isinstance(val, str):
        cleaned = val.strip()
        if cleaned == "" or cleaned.lower() in ("nan", "null", "none"):
            return None
        return cleaned
    return val

def parse_csv_value(val_str):
    """Parse CSV text value into appropriate type (int, float, or string)."""
    cleaned = clean_cell_value(val_str)
    if cleaned is None:
        return None
    if not isinstance(cleaned, str):
        return cleaned

    # Try integer
    try:
        val_int = int(cleaned)
        return val_int
    except ValueError:
        pass

    # Try float
    try:
        val_float = float(cleaned)
        if math.isnan(val_float):
            return None
        return val_float
    except ValueError:
        pass

    return cleaned

def resolve_header_grid(ws: Worksheet, header_rows: int = 2) -> list[tuple[str, str]]:
    """
    Constructs a 2-level header tuple for each column in the worksheet.
    Resolves merged cell ranges so merged parent titles are copied across all spanned columns.
    Returns list of tuples: [(header_row1, header_row2), ...] for each column index.
    """
    max_col = ws.max_column
    if max_col < 1:
        return []

    # Initialize grid [row][col] (1-indexed)
    grid = {}
    for r in range(1, header_rows + 1):
        grid[r] = {}
        for c in range(1, max_col + 1):
            grid[r][c] = clean_cell_value(ws.cell(row=r, column=c).value)

    # Resolve merged cell ranges in the header rows
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col_m, max_row_m = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )

        if min_row <= header_rows:
            top_left_val = clean_cell_value(ws.cell(row=min_row, column=min_col).value)
            end_r = min(max_row_m, header_rows)
            for r in range(min_row, end_r + 1):
                for c in range(min_col, max_col_m + 1):
                    if grid[r][c] is None and top_left_val is not None:
                        grid[r][c] = top_left_val

    # Forward fill top row if any remaining gaps from CSV-style merged export
    last_top = None
    for c in range(1, max_col + 1):
        curr_top = grid[1].get(c)
        if curr_top is not None:
            last_top = curr_top
        elif last_top is not None and (grid[2].get(c) is not None):
            grid[1][c] = last_top

    # Construct list of header tuples for each column (1..max_col)
    col_headers = []
    for c in range(1, max_col + 1):
        top_h = grid[1].get(c)
        bot_h = grid[2].get(c) if header_rows >= 2 else None
        
        top_str = str(top_h) if top_h is not None else ""
        bot_str = str(bot_h) if bot_h is not None else ""
        
        col_headers.append((top_str, bot_str))

    return col_headers

class ExcelReader:
    def __init__(self, file_path: str, batch_filter_pattern: str = r"^B\d+P\d+", header_rows: int = 2):
        self.file_path = Path(file_path)
        self.batch_regex = re.compile(batch_filter_pattern, re.IGNORECASE)
        self.header_rows = header_rows

    def is_csv(self) -> bool:
        return self.file_path.suffix.lower() == ".csv"

    def get_matching_sheets(self, wb) -> list[str]:
        matching = []
        for name in wb.sheetnames:
            if self.batch_regex.search(name.strip()):
                matching.append(name)
        return matching

    def read_csv(self):
        """Reads a CSV file supporting multi-row headers and yields row tuples."""
        logger.info("Opening CSV file: %s", self.file_path)

        # Try reading with utf-8-sig (handles BOM), fallback to latin-1
        lines = []
        try:
            with open(self.file_path, mode="r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                lines = list(reader)
        except UnicodeDecodeError:
            with open(self.file_path, mode="r", encoding="latin-1") as f:
                reader = csv.reader(f)
                lines = list(reader)

        if len(lines) <= self.header_rows:
            logger.warning("CSV file %s has fewer rows (%d) than header rows (%d)",
                           self.file_path, len(lines), self.header_rows)
            return

        row1_raw = lines[0] if len(lines) > 0 else []
        row2_raw = lines[1] if self.header_rows >= 2 and len(lines) > 1 else []

        max_col = max(len(row1_raw), len(row2_raw))

        # Build col_headers with forward-fill for Row 1 merged headers
        top_headers = []
        last_top = ""
        for c in range(max_col):
            t_val = clean_cell_value(row1_raw[c]) if c < len(row1_raw) else None
            b_val = clean_cell_value(row2_raw[c]) if c < len(row2_raw) else None

            if t_val is not None:
                last_top = str(t_val)
            elif last_top and b_val is not None:
                t_val = last_top

            top_headers.append((str(t_val) if t_val else "", str(b_val) if b_val else ""))

        sheet_name = self.file_path.stem
        logger.info("Processing CSV '%s' (%d data rows, %d columns)", sheet_name, len(lines) - self.header_rows, len(top_headers))

        for r_idx in range(self.header_rows, len(lines)):
            row_data = lines[r_idx]
            row_raw = {}
            row_list = []
            is_empty_row = True

            for c_idx, (top_h, bot_h) in enumerate(top_headers):
                raw_cell = row_data[c_idx] if c_idx < len(row_data) else None
                val = parse_csv_value(raw_cell)
                if val is not None:
                    is_empty_row = False
                row_raw[(top_h, bot_h)] = val
                row_list.append(((top_h, bot_h), val))

            if is_empty_row:
                continue

            yield sheet_name, r_idx + 1, row_raw, row_list

    def read_workbook(self):
        """
        Yields (sheet_name, row_idx, row_values_dict, row_list) for matching Excel sheets or CSV file.
        """
        if self.is_csv():
            yield from self.read_csv()
            return

        logger.info("Opening Excel workbook: %s", self.file_path)
        wb = openpyxl.load_workbook(filename=str(self.file_path), data_only=True)
        matching_sheets = self.get_matching_sheets(wb)

        if not matching_sheets:
            logger.warning("No sheets matched regex pattern '%s' in workbook %s. Found sheet names: %s",
                           self.batch_regex.pattern, self.file_path, wb.sheetnames)
            return

        logger.info("Matched %d sheets for processing: %s", len(matching_sheets), matching_sheets)

        for sheet_name in matching_sheets:
            ws = wb[sheet_name]
            col_headers = resolve_header_grid(ws, header_rows=self.header_rows)
            max_row = ws.max_row

            logger.info("Processing sheet '%s' (%d total rows, %d columns)", sheet_name, max_row, len(col_headers))

            for r in range(self.header_rows + 1, max_row + 1):
                row_raw = {}
                row_list = []
                is_empty_row = True

                for c_idx, (top_h, bot_h) in enumerate(col_headers, start=1):
                    val = clean_cell_value(ws.cell(row=r, column=c_idx).value)
                    if val is not None:
                        is_empty_row = False
                    row_raw[(top_h, bot_h)] = val
                    row_list.append(((top_h, bot_h), val))

                if is_empty_row:
                    continue

                yield sheet_name, r, row_raw, row_list
